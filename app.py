#!/usr/bin/env python
# coding: utf-8

# In[1]:


import io
import json
import os
import pandas as pd
import streamlit as st
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from Levenshtein import distance as levenshtein_distance
import re
from openai import OpenAI
import random
import time

# Set up the OpenAI client with error handling
try:
    client = OpenAI(api_key=st.secrets["OPENAI_API_KEY"])
except KeyError:
    st.error("OpenAI API key not found in secrets. Please check your configuration.")
    st.stop()

# Function to generate a response from OpenAI
def generate_response(prompt):
    try:
        response = client.chat.completions.create(
            model="gpt-4o",
            messages=[{"role": "user", "content": prompt}],
            max_tokens=1000,
            temperature=0.2
        )
        return response.choices[0].message.content.strip()
    except Exception as e:
        st.error(f"An error occurred: {str(e)}")
        return "I'm sorry, but I encountered an error while processing your request."

# Function to get AI-suggested mapping
def get_ai_suggested_mapping(label, account, balance_sheet_lookup_df):
    prompt = f"""Given the following account information:
    Label: {label}
    Account: {account}

    And the following balance sheet lookup data:
    {balance_sheet_lookup_df.to_string()}

    What is the most appropriate Mnemonic mapping for this account based on Label and Account combination? Please provide only the value from the 'Mnemonic' column in the Balance Sheet Data Dictionary data frame based on Label and Account combination, without any explanation. The determination should be based on business logic first then similarity. Ensure that the suggested Mnemonic is appropriate for the given Label e.g., don't suggest a Current Asset Mnemonic for a current liability Label."""

    suggested_mnemonic = generate_response(prompt).strip()


    # Check if the suggested_mnemonic is in the Mnemonic column
    if suggested_mnemonic in balance_sheet_lookup_df['Mnemonic'].values:
        return suggested_mnemonic
    else:
        # If not, try to find a matching row based on Label and Account
        matching_row = balance_sheet_lookup_df[
            (balance_sheet_lookup_df['Label'].str.lower() == label.lower()) &
            (balance_sheet_lookup_df['Account'].str.lower() == account.lower())
        ]
        if not matching_row.empty:
            return matching_row['Mnemonic'].values[0]
        else:
            # If still no match, find the closest match based on Levenshtein distance
            best_match = None
            best_score = float('inf')
            for _, row in balance_sheet_lookup_df.iterrows():
                if row['Label'].strip().lower() == label.strip().lower():
                    score = levenshtein_distance(account.lower(), row['Account'].lower())
                    if score < best_score:
                        best_score = score
                        best_match = row['Mnemonic']
            
            if best_match:
                return f"{best_match}"
            else:
                return "No matching Mnemonic found"

# Define the initial lookup data for Balance Sheet
initial_balance_sheet_lookup_data = {
    "Label": ["Current Assets", "Non Current Assets", "Non Current Assets", "Non Current Assets", "Non Current Assets", "Non Current Assets", "Non Current Assets"],
    "Account": ["Gross Property, Plant & Equipment", "Accumulated Depreciation", "Net Property, Plant & Equipment", "Long-term Investments", "Goodwill", "Other Intangibles", "Right-of-Use Asset-Net"],
    "Mnemonic": ["Gross Property, Plant & Equipment", "Accumulated Depreciation", "Net Property, Plant & Equipment", "Long-term Investments", "Goodwill", "Other Intangibles", "Right-of-Use Asset-Net"],
    "CIQ": ["IQ_GPPE", "IQ_AD", "IQ_NPPE", "IQ_LT_INVEST", "IQ_GW", "IQ_OTHER_INTAN", "IQ_RUA_NET"]
}

# Define the file paths for the data dictionaries
balance_sheet_data_dictionary_file = 'balance_sheet_data_dictionary.csv'

# Load or initialize the lookup table
def load_or_initialize_lookup(file_path, initial_data):
    if os.path.exists(file_path):
        lookup_df = pd.read_csv(file_path)
    else:
        lookup_df = pd.DataFrame(initial_data)
        lookup_df.to_csv(file_path, index=False)
    return lookup_df

def save_lookup_table_bs_cf(df, file_path):
    df.to_csv(file_path, index=False)

# Initialize lookup tables for Balance Sheet
balance_sheet_lookup_df = load_or_initialize_lookup(balance_sheet_data_dictionary_file, initial_balance_sheet_lookup_data)

# General Utility Functions
def process_file(file):
    try:
        df = pd.read_excel(file, sheet_name=None)
        first_sheet_name = list(df.keys())[0]
        df = df[first_sheet_name]
        return df
    except Exception as e:
        st.error(f"Error processing file {file.name}: {e}")
        return None

def create_combined_df(dfs):
    combined_df = pd.DataFrame()
    for i, df in enumerate(dfs):
        final_mnemonic_col = 'Final Mnemonic Selection'
        if final_mnemonic_col not in df.columns:
            st.error(f"Column '{final_mnemonic_col}' not found in dataframe {i+1}")
            continue
        
        date_cols = [col for col in df.columns if col not in ['Label', 'Account', final_mnemonic_col, 'Mnemonic', 'Manual Selection']]
        if not date_cols:
            st.error(f"No date columns found in dataframe {i+1}")
            continue

        df_grouped = df.groupby([final_mnemonic_col, 'Label']).sum(numeric_only=True).reset_index()
        df_melted = df_grouped.melt(id_vars=[final_mnemonic_col, 'Label'], value_vars=date_cols, var_name='Date', value_name='Value')
        df_pivot = df_melted.pivot(index=['Label', final_mnemonic_col], columns='Date', values='Value')
        
        if combined_df.empty:
            combined_df = df_pivot
        else:
            combined_df = combined_df.join(df_pivot, how='outer')
    return combined_df.reset_index()

def aggregate_data(df):
    if 'Label' not in df.columns or 'Account' not in df.columns:
        st.error("'Label' and/or 'Account' columns not found in the data.")
        return df
    
    pivot_table = df.pivot_table(index=['Label', 'Account'], 
                                 values=[col for col in df.columns if col not in ['Label', 'Account', 'Mnemonic', 'Manual Selection']], 
                                 aggfunc='sum').reset_index()
    return pivot_table

def clean_numeric_value(value):
    value_str = str(value).strip()
    if value_str.startswith('(') and value_str.endswith(')'):
        value_str = '-' + value_str[1:-1]
    cleaned_value = re.sub(r'[$,]', '', value_str)
    try:
        return float(cleaned_value)
    except ValueError:
        return 0

def sort_by_label_and_account(df):
    sort_order = {
        "Current Assets": 0,
        "Non Current Assets": 1,
        "Current Liabilities": 2,
        "Non Current Liabilities": 3,
        "Equity": 4,
        "Total Equity and Liabilities": 5
    }
    
    df['Label_Order'] = df['Label'].map(sort_order)
    df['Total_Order'] = df['Account'].str.contains('Total', case=False).astype(int)
    
    df = df.sort_values(by=['Label_Order', 'Label', 'Total_Order', 'Account']).drop(columns=['Label_Order', 'Total_Order'])
    return df

def sort_by_label_and_final_mnemonic(df):
    sort_order = {
        "Current Assets": 0,
        "Non Current Assets": 1,
        "Current Liabilities": 2,
        "Non Current Liabilities": 3,
        "Equity": 4,
        "Total Equity and Liabilities": 5
    }
    
    df['Label_Order'] = df['Label'].map(sort_order)
    df['Total_Order'] = df['Final Mnemonic Selection'].str.contains('Total', case=False).astype(int)
    
    df = df.sort_values(by=['Label_Order', 'Total_Order', 'Final Mnemonic Selection']).drop(columns=['Label_Order', 'Total_Order'])
    return df

def apply_unit_conversion(df, columns, factor):
    for selected_column in columns:
        if selected_column in df.columns:
            df[selected_column] = df[selected_column].apply(
                lambda x: x * factor if isinstance(x, (int, float)) else x)
    return df

# Function to check if all values in columns past the first 2 columns are 0
def check_all_zeroes(df):
    zeroes = (df.iloc[:, 2:] == 0).all(axis=1)
    return zeroes

# Balance Sheet Functions
def balance_sheet():
    global balance_sheet_lookup_df

    st.title("BALANCE SHEET LTMA")

    tab1, tab2, tab3, tab4 = st.tabs(["Table Extractor", "Aggregate My Data", "Mappings and Data Consolidation", "Balance Sheet Data Dictionary"])

    with tab1:
        uploaded_file = st.file_uploader("Choose a JSON file", type="json", key='json_uploader')
        if uploaded_file is not None:
            data = json.load(uploaded_file)
            st.warning("PLEASE NOTE: In the Setting Bounds Preview Window, you will see only your respective labels. In the Updated Columns Preview Window, you will see only your renamed column headers. The labels from the Setting Bounds section will not appear in the Updated Columns Preview.")
            st.warning("PLEASE ALSO NOTE: An Account column must also be designated when you are in the Rename Columns section.")

            tables = []
            for block in data['Blocks']:
                if block['BlockType'] == 'TABLE':
                    table = {}
                    if 'Relationships' in block:
                        for relationship in block['Relationships']:
                            if relationship['Type'] == 'CHILD':
                                for cell_id in relationship['Ids']:
                                    cell_block = next((b for b in data['Blocks'] if b['Id'] == cell_id), None)
                                    if cell_block:
                                        row_index = cell_block.get('RowIndex', 0)
                                        col_index = cell_block.get('ColumnIndex', 0)
                                        if row_index not in table:
                                            table[row_index] = {}
                                        cell_text = ''
                                        if 'Relationships' in cell_block:
                                            for rel in cell_block['Relationships']:
                                                if rel['Type'] == 'CHILD':
                                                    for word_id in rel['Ids']:
                                                        word_block = next((w for w in data['Blocks'] if w['Id'] == word_id), None)
                                                        if word_block and word_block['BlockType'] == 'WORD':
                                                            cell_text += ' ' + word_block.get('Text', '')
                                        table[row_index][col_index] = cell_text.strip()
                    table_df = pd.DataFrame.from_dict(table, orient='index').sort_index()
                    table_df = table_df.sort_index(axis=1)
                    tables.append(table_df)
            all_tables = pd.concat(tables, axis=0, ignore_index=True)
            if len(all_tables.columns) == 0:
                st.error("No columns found in the uploaded JSON file.")
                return

            column_a = all_tables.columns[0]
            all_tables.insert(0, 'Label', '')

            st.subheader("Data Preview")
            st.dataframe(all_tables)

            def get_unique_options(series):
                counts = series.value_counts()
                unique_options = []
                occurrence_counts = {}
                for item in series:
                    if counts[item] > 1:
                        if item not in occurrence_counts:
                            occurrence_counts[item] = 1
                        else:
                            occurrence_counts[item] += 1
                        unique_options.append(f"{item} {occurrence_counts[item]}")
                    else:
                        unique_options.append(item)
                return unique_options

            labels = ["Current Assets", "Non Current Assets", "Current Liabilities",
                      "Non Current Liabilities", "Equity", "Total Equity and Liabilities"]
            selections = []

            for label in labels:
                st.subheader(f"Setting bounds for {label}")
                options = [''] + get_unique_options(all_tables[column_a].dropna())
                start_label = st.selectbox(f"Start Label for {label}", options, key=f"start_{label}")
                end_label = st.selectbox(f"End Label for {label}", options, key=f"end_{label}")
                selections.append((label, start_label, end_label))

            new_column_names = {col: col for col in all_tables.columns}

            def update_labels(df):
                df['Label'] = ''
                account_column = new_column_names.get(column_a, column_a)
                for label, start_label, end_label in selections:
                    if start_label and end_label:
                        try:
                            start_label_base = " ".join(start_label.split()[:-1]) if start_label.split()[-1].isdigit() else start_label
                            end_label_base = " ".join(end_label.split()[:-1]) if end_label.split()[-1].isdigit() else end_label

                            start_index = df[df[account_column] == start_label_base].index.min()
                            end_index = df[df[account_column] == end_label_base].index.max()

                            if pd.isna(start_index):
                                start_index = df[df[account_column].str.contains(start_label_base, regex=False, na=False)].index.min()
                            if pd.isna(end_index):
                                end_index = df[df[account_column].str.contains(end_label_base, regex=False, na=False)].index.max()

                            if pd.notna(start_index) and pd.notna(end_index):
                                df.loc[start_index:end_index, 'Label'] = label
                            else:
                                st.error(f"Invalid label bounds for {label}. Skipping...")
                        except KeyError as e:
                            st.error(f"Error accessing column '{account_column}': {e}. Skipping...")
                    else:
                        st.info(f"No selections made for {label}. Skipping...")
                return df

            if st.button("Preview Setting Bounds ONLY", key="preview_setting_bounds"):
                preview_table = update_labels(all_tables.copy())
                st.subheader("Preview of Setting Bounds")
                st.dataframe(preview_table)

            st.subheader("Rename Columns")
            new_column_names = {}
            fiscal_year_options = [f"FY{year}" for year in range(2018, 2027)]
            ytd_options = [f"YTD{quarter}{year}" for year in range(2018, 2027) for quarter in range(1, 4)]
            dropdown_options = [''] + ['Account'] + fiscal_year_options + ytd_options

            for col in all_tables.columns:
                new_name_text = st.text_input(f"Rename '{col}' to:", value=col, key=f"rename_{col}_text")
                new_name_dropdown = st.selectbox(f"Or select predefined name for '{col}':", dropdown_options, key=f"rename_{col}_dropdown", index=0)
                new_column_names[col] = new_name_dropdown if new_name_dropdown else new_name_text

            all_tables.rename(columns=new_column_names, inplace=True)
            st.write("Updated Columns:", all_tables.columns.tolist())
            st.dataframe(all_tables)

            st.subheader("Select columns to keep before export")
            columns_to_keep = []
            for col in all_tables.columns:
                if st.checkbox(f"Keep column '{col}'", value=True, key=f"keep_{col}"):
                    columns_to_keep.append(col)

            st.subheader("Select numerical columns")
            numerical_columns = []
            for col in all_tables.columns:
                if st.checkbox(f"Numerical column '{col}'", value=False, key=f"num_{col}"):
                    numerical_columns.append(col)

            if 'Label' not in columns_to_keep:
                columns_to_keep.insert(0, 'Label')

            if 'Account' not in columns_to_keep:
                columns_to_keep.insert(1, 'Account')

            st.subheader("Label Units")
            selected_columns = st.multiselect("Select columns for conversion", options=numerical_columns, key="columns_selection")
            selected_value = st.radio("Select conversion value", ["Actuals", "Thousands", "Millions", "Billions"], index=0, key="conversion_value")

            conversion_factors = {
                "Actuals": 1,
                "Thousands": 1000,
                "Millions": 1000000,
                "Billions": 1000000000
            }

            if st.button("Apply Selected Labels and Generate Excel", key="apply_selected_labels_generate_excel_tab1"):
                updated_table = update_labels(all_tables.copy())
                updated_table = updated_table[[col for col in columns_to_keep if col in updated_table.columns]]

                updated_table = updated_table[updated_table['Label'].str.strip() != '']
                updated_table = updated_table[updated_table['Account'].str.strip() != '']

                for col in numerical_columns:
                    if col in updated_table.columns:
                        updated_table[col] = updated_table[col].apply(clean_numeric_value)

                if selected_value != "Actuals":
                    updated_table = apply_unit_conversion(updated_table, selected_columns, conversion_factors[selected_value])

                updated_table.replace('-', 0, inplace=True)

                excel_file = io.BytesIO()
                updated_table.to_excel(excel_file, index=False)
                excel_file.seek(0)
                st.download_button("Download Excel", excel_file, "Table_Extractor_Balance_Sheet.xlsx", "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

            st.subheader("Check for Duplicate Accounts")
            if 'Account' not in all_tables.columns:
                st.warning("The 'Account' column is missing. Please ensure your data includes an 'Account' column.")
            else:
                duplicated_accounts = all_tables[all_tables.duplicated(['Account'], keep=False)]
                if not duplicated_accounts.empty:
                    st.warning("Duplicates identified:")
                    st.dataframe(duplicated_accounts)
                else:
                    st.success("No duplicates identified")

    with tab2:
        st.subheader("Aggregate My Data")

        uploaded_files = st.file_uploader("Upload your Excel files from Tab 1", type=['xlsx'], accept_multiple_files=True, key='xlsx_uploader_tab2')

        dfs = []
        if uploaded_files:
            dfs = [process_file(file) for file in uploaded_files if process_file(file) is not None]

        if dfs:
            combined_df = pd.concat(dfs, ignore_index=True)
            st.dataframe(combined_df)

            aggregated_table = aggregate_data(combined_df)
            aggregated_table = sort_by_label_and_account(aggregated_table)

            st.subheader("Aggregated Data")
            st.dataframe(aggregated_table)

            st.subheader("Preview Data and Edit Rows")
            zero_rows = check_all_zeroes(aggregated_table)  # Check for rows with all zero values
            zero_rows_indices = aggregated_table.index[zero_rows].tolist()
            st.write("Rows where all values (past the first 2 columns) are zero:", aggregated_table.loc[zero_rows_indices])

            edited_data = st.experimental_data_editor(aggregated_table, num_rows="dynamic")

            # Highlight rows with all zeros for potential removal
            st.write("Highlighted rows with all zero values for potential removal:")
            for index in zero_rows_indices:
                st.write(f"Row {index}: {aggregated_table.loc[index].to_dict()}")

            rows_removed = False  # Flag to check if rows are removed
            if st.button("Remove Highlighted Rows", key="remove_highlighted_rows"):
                aggregated_table = aggregated_table.drop(zero_rows_indices).reset_index(drop=True)
                rows_removed = True
                st.success("Highlighted rows removed successfully")
                st.dataframe(aggregated_table)

            st.subheader("Download Aggregated Data")
            if rows_removed:
                download_label = "Download Updated Aggregated Excel"
            else:
                download_label = "Download Aggregated Excel"
            excel_file = io.BytesIO()
            with pd.ExcelWriter(excel_file, engine='xlsxwriter') as writer:
                aggregated_table.to_excel(writer, sheet_name='Aggregated Data', index=False)
            excel_file.seek(0)
            st.download_button(download_label, excel_file, "Aggregate_My_Data_Balance_Sheet.xlsx", "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        else:
            st.warning("Please upload valid Excel files for aggregation.")

    with tab3:
        st.subheader("Mappings and Data Consolidation")

        uploaded_excel = st.file_uploader("Upload your Excel file for Mnemonic Mapping", type=['xlsx'], key='excel_uploader_tab3_bs')

        currency_options = ["U.S. Dollar", "Euro", "British Pound Sterling", "Japanese Yen"]
        magnitude_options = ["Actuals", "Thousands", "Millions", "Billions", "Trillions"]

        selected_currency = st.selectbox("Select Currency", currency_options, key='currency_selection_tab3_bs')
        selected_magnitude = st.selectbox("Select Magnitude", magnitude_options, key='magnitude_selection_tab3_bs')
        company_name_bs = st.text_input("Enter Company Name", key='company_name_input_bs')

        if uploaded_excel is not None:
            df = pd.read_excel(uploaded_excel)

            statement_dates = {}
            for col in df.columns[2:]:
                statement_date = st.text_input(f"Enter statement date for {col}", key=f"statement_date_{col}")
                statement_dates[col] = statement_date

            st.write("Columns in the uploaded file:", df.columns.tolist())

            if 'Account' not in df.columns:
                st.error("The uploaded file does not contain an 'Account' column.")
            else:
                # Function to get the best match based on Label first, then Levenshtein distance on Account
                def get_best_match(label, account):
                    best_score = float('inf')
                    best_match = None
                    for _, lookup_row in balance_sheet_lookup_df.iterrows():
                        if 'Label' in lookup_row and lookup_row['Label'].strip().lower() == str(label).strip().lower():
                            lookup_account = lookup_row['Account']
                            account_str = str(account)
                            # Levenshtein distance for Account
                            score = levenshtein_distance(account_str.lower(), lookup_account.lower()) / max(len(account_str), len(lookup_account))
                            if score < 0.25 and score < best_score:
                                best_score = score
                                best_match = lookup_row
                    return best_match, best_score

                df['Mnemonic'] = ''
                df['Manual Selection'] = ''
                for idx, row in df.iterrows():
                    account_value = row['Account']
                    label_value = row.get('Label', '')
                    if pd.notna(account_value):
                        best_match, score = get_best_match(label_value, account_value)
                        if best_match is not None:
                            df.at[idx, 'Mnemonic'] = best_match['Mnemonic']
                        else:
                            df.at[idx, 'Mnemonic'] = 'Human Intervention Required'
                            if f"ai_called_{idx}" not in st.session_state:
                                ai_suggested_mnemonic = get_ai_suggested_mapping(label_value, account_value, balance_sheet_lookup_df)
                                st.session_state[f"ai_called_{idx}"] = ai_suggested_mnemonic
                                st.markdown(f"**Human Intervention Required for:** {account_value} [{label_value} - Index {idx}]")
                                st.markdown(f"**AI Suggested Mapping:** {ai_suggested_mnemonic}")
                            else:
                                ai_suggested_mnemonic = st.session_state[f"ai_called_{idx}"]
                                st.markdown(f"**Human Intervention Required for:** {account_value} [{label_value} - Index {idx}]")
                                st.markdown(f"**AI Suggested Mapping:** {ai_suggested_mnemonic}")

                    # Create a dropdown list of unique mnemonics based on the label
                    label_mnemonics = balance_sheet_lookup_df[balance_sheet_lookup_df['Label'] == label_value]['Mnemonic'].unique()
                    manual_selection_options = [mnemonic for mnemonic in label_mnemonics]
                    manual_selection = st.selectbox(
                        f"Select category for '{account_value}'",
                        options=[''] + manual_selection_options + ['REMOVE ROW'],
                        key=f"select_{idx}_tab3_bs"
                    )
                    if manual_selection:
                        df.at[idx, 'Manual Selection'] = manual_selection.strip()

                st.dataframe(df[['Label', 'Account', 'Mnemonic', 'Manual Selection']])

                if st.button("Generate Excel with Lookup Results", key="generate_excel_lookup_results_tab3_bs"):
                    df['Final Mnemonic Selection'] = df.apply(
                        lambda row: row['Manual Selection'].strip() if row['Manual Selection'].strip() != '' else row['Mnemonic'],
                        axis=1
                    )
                    final_output_df = df[df['Final Mnemonic Selection'].str.strip() != 'REMOVE ROW'].copy()

                    combined_df = create_combined_df([final_output_df])
                    combined_df = sort_by_label_and_final_mnemonic(combined_df)

                    columns_order = ['Label', 'Final Mnemonic Selection'] + [col for col in combined_df.columns if col not in ['Label', 'Final Mnemonic Selection']]
                    combined_df = combined_df[columns_order]

                    # Include the "As Presented" sheet without the CIQ column, and with the specified column order
                    as_presented_df = final_output_df.drop(columns=['CIQ', 'Mnemonic', 'Manual Selection'], errors='ignore')
                    as_presented_columns_order = ['Label', 'Account', 'Final Mnemonic Selection'] + [col for col in as_presented_df.columns if col not in ['Label', 'Account', 'Final Mnemonic Selection']]
                    as_presented_df = as_presented_df[as_presented_columns_order]

                    excel_file = io.BytesIO()
                    with pd.ExcelWriter(excel_file, engine='xlsxwriter') as writer:
                        combined_df.to_excel(writer, sheet_name='Standardized - Balance Sheet', index=False)
                        as_presented_df.to_excel(writer, sheet_name='As Presented - Balance Sheet', index=False)
                        cover_df = pd.DataFrame({
                            'Selection': ['Currency', 'Magnitude', 'Company Name'] + list(statement_dates.keys()),
                            'Value': [selected_currency, selected_magnitude, company_name_bs] + list(statement_dates.values())
                        })
                        cover_df.to_excel(writer, sheet_name='Cover', index=False)
                    excel_file.seek(0)
                    st.download_button("Download Excel", excel_file, "Mappings_and_Data_Consolidation_Balance_Sheet.xlsx", "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

                if st.button("Update Data Dictionary with Manual Mappings", key="update_data_dictionary_tab3_bs"):
                    df['Final Mnemonic Selection'] = df.apply(
                        lambda row: row['Manual Selection'] if row['Manual Selection'] not in ['REMOVE ROW', ''] else row['Mnemonic'],
                        axis=1
                    )
                    new_entries = []
                    for idx, row in df.iterrows():
                        manual_selection = row['Manual Selection']
                        final_mnemonic = row['Final Mnemonic Selection']
                        if manual_selection == 'REMOVE ROW':
                            continue

                        if manual_selection not in ['REMOVE ROW', '']:
                            if row['Account'] not in balance_sheet_lookup_df['Account'].values:
                                new_entries.append({'Account': row['Account'], 'Mnemonic': final_mnemonic, 'Label': row['Label']})
                            else:
                                balance_sheet_lookup_df.loc[balance_sheet_lookup_df['Account'] == row['Account'], 'Mnemonic'] = final_mnemonic
                                balance_sheet_lookup_df.loc[balance_sheet_lookup_df['Account'] == row['Account'], 'Label'] = row['Label']
                    if new_entries:
                        balance_sheet_lookup_df = pd.concat([balance_sheet_lookup_df, pd.DataFrame(new_entries)], ignore_index=True)
                    balance_sheet_lookup_df.reset_index(drop=True, inplace=True)
                    save_lookup_table_bs_cf(balance_sheet_lookup_df, balance_sheet_data_dictionary_file)
                    st.success("Data Dictionary Updated Successfully")


    with tab4:
        st.subheader("Balance Sheet Data Dictionary")

        uploaded_dict_file = st.file_uploader("Upload a new Data Dictionary CSV", type=['csv'], key='dict_uploader_tab4_bs')
        if uploaded_dict_file is not None:
            new_lookup_df = pd.read_csv(uploaded_dict_file)
            balance_sheet_lookup_df = new_lookup_df  # Overwrite the entire DataFrame
            save_lookup_table_bs_cf(balance_sheet_lookup_df, balance_sheet_data_dictionary_file)
            st.success("Data Dictionary uploaded and updated successfully!")

        st.dataframe(balance_sheet_lookup_df)

        remove_indices = st.multiselect("Select rows to remove", balance_sheet_lookup_df.index, key='remove_indices_tab4_bs')
        rows_removed = False
        if st.button("Remove Selected Rows", key="remove_selected_rows_tab4_bs"):
            balance_sheet_lookup_df = balance_sheet_lookup_df.drop(remove_indices).reset_index(drop=True)
            save_lookup_table_bs_cf(balance_sheet_lookup_df, balance_sheet_data_dictionary_file)
            rows_removed = True
            st.success("Selected rows removed successfully!")
            st.dataframe(balance_sheet_lookup_df)

        st.subheader("Download Data Dictionary")
        if rows_removed:
            download_label = "Download Updated Data Dictionary"
        else:
            download_label = "Download Data Dictionary"
        excel_file = io.BytesIO()
        balance_sheet_lookup_df.to_excel(excel_file, index=False)
        excel_file.seek(0)
        st.download_button(download_label, excel_file, "balance_sheet_data_dictionary.xlsx", "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

#############################INCOME STATEMENT#######################################################################
import io
import os
import re
import json
import pandas as pd
import streamlit as st
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from Levenshtein import distance as levenshtein_distance

# Ensure conversion_factors is defined
conversion_factors = {
    "Actuals": 1,
    "Thousands": 1000,
    "Millions": 1000000,
    "Billions": 1000000000
}

def clean_numeric_value_IS(value):
    try:
        value_str = str(value).strip()
        if value_str.startswith('(') and value_str.endswith(')'):
            value_str = '-' + value_str[1:-1]
        cleaned_value = re.sub(r'[$,]', '', value_str)
        return float(cleaned_value)
    except (ValueError, TypeError):
        return value

def apply_unit_conversion_IS(df, columns, factor):
    for selected_column in columns:
        if selected_column in df.columns:
            df[selected_column] = df[selected_column].apply(
                lambda x: x * factor if isinstance(x, (int, float)) else x)
    return df

def create_combined_df_IS(dfs):
    combined_df = pd.DataFrame()
    for i, df in enumerate(dfs):
        final_mnemonic_col = 'Final Mnemonic Selection'
        if final_mnemonic_col not in df.columns:
            st.error(f"Column '{final_mnemonic_col}' not found in dataframe {i+1}")
            continue
        
        date_cols = [col for col in df.columns if col not in ['Account', final_mnemonic_col, 'Mnemonic', 'Manual Selection', 'Sort Index']]
        if not date_cols:
            st.error(f"No date columns found in dataframe {i+1}")
            continue

        df_grouped = df.groupby([final_mnemonic_col]).sum(numeric_only=True).reset_index()
        df_melted = df_grouped.melt(id_vars=[final_mnemonic_col], value_vars=date_cols, var_name='Date', value_name='Value')
        df_pivot = df_melted.pivot(index=[final_mnemonic_col], columns='Date', values='Value')
        
        # Reverse the order of the date columns
        df_pivot = df_pivot[sorted(df_pivot.columns, reverse=True)]
        
        if combined_df.empty:
            combined_df = df_pivot
        else:
            combined_df = combined_df.join(df_pivot, how='outer')
    return combined_df.reset_index()

def sort_by_sort_index(df):
    if 'Sort Index' in df.columns:
        df = df.sort_values(by=['Sort Index'])
    return df

def aggregate_data_IS(uploaded_files):
    dataframes = []
    unique_accounts = set()

    for file in uploaded_files:
        df = pd.read_excel(file)
        df.columns = [str(col).strip() for col in df.columns]

        if 'Account' not in df.columns:
            st.error(f"Column 'Account' not found in file {file.name}")
            return None

        df['Sort Index'] = range(1, len(df) + 1)
        dataframes.append(df)
        unique_accounts.update(df['Account'].dropna().unique())

    concatenated_df = pd.concat(dataframes, ignore_index=True)

    statement_date_rows = concatenated_df[concatenated_df['Account'].str.contains('Statement Date:', na=False)]
    numeric_rows = concatenated_df[~concatenated_df['Account'].str.contains('Statement Date:', na=False)]

    for col in numeric_rows.columns:
        if col not in ['Account', 'Sort Index', 'Positive Decreases NI']:
            numeric_rows[col] = numeric_rows[col].apply(clean_numeric_value_IS)

    numeric_rows.fillna(0, inplace=True)

    for col in numeric_rows.columns:
        if col not in ['Account', 'Sort Index', 'Positive Decreases NI']:
            numeric_rows[col] = pd.to_numeric(numeric_rows[col], errors='coerce').fillna(0)

    aggregated_df = numeric_rows.groupby(['Account'], as_index=False).sum(min_count=1)

    statement_date_rows['Sort Index'] = 100
    statement_date_rows = statement_date_rows.groupby('Account', as_index=False).first()

    final_df = pd.concat([aggregated_df, statement_date_rows], ignore_index=True)

    final_df.insert(1, 'Positive Decreases NI', False)

    sort_index_column = final_df.pop('Sort Index')
    final_df['Sort Index'] = sort_index_column

    final_df.sort_values('Sort Index', inplace=True)

    return final_df

def save_lookup_table(df, file_path):
    df.to_excel(file_path, index=False)

def update_negative_values(df):
    criteria = [
        "IQ_COGS",
        "IQ_SGA_SUPPL",
        "IQ_RD_EXP",
        "IQ_DA_SUPPL",
        "IQ_STOCK_BASED",
        "IQ_OTHER_OPER",
        "IQ_INC_TAX"
    ]
    
    for index, row in df.iterrows():
        if row['CIQ'] in criteria:
            for col in df.columns[2:]:  # Start from column index 2 (skip 'Final Mnemonic Selection' and 'CIQ')
                if isinstance(row[col], (int, float)) and row[col] < 0:
                    df.at[index, col] = row[col] * -1
    return df

def income_statement():
    global income_statement_lookup_df

    if 'income_statement_lookup_df' not in globals():
        if os.path.exists(income_statement_data_dictionary_file):
            income_statement_lookup_df = pd.read_excel(income_statement_data_dictionary_file)
        else:
            income_statement_lookup_df = pd.DataFrame()

    st.title("INCOME STATEMENT LTMA")
    tab1, tab2, tab3, tab4 = st.tabs(["Table Extractor", "Aggregate My Data", "Mappings and Data Consolidation", "Income Statement Data Dictionary"])

    with tab1:
        uploaded_file = st.file_uploader("Choose a JSON file", type="json", key='json_uploader')
        if uploaded_file is not None:
            data = json.load(uploaded_file)
            tables = []
            for block in data['Blocks']:
                if block['BlockType'] == 'TABLE':
                    table = {}
                    if 'Relationships' in block:
                        for relationship in block['Relationships']:
                            if relationship['Type'] == 'CHILD':
                                for cell_id in relationship['Ids']:
                                    cell_block = next((b for b in data['Blocks'] if b['Id'] == cell_id), None)
                                    if cell_block:
                                        row_index = cell_block.get('RowIndex', 0)
                                        col_index = cell_block.get('ColumnIndex', 0)
                                        if row_index not in table:
                                            table[row_index] = {}
                                        cell_text = ''
                                        if 'Relationships' in cell_block:
                                            for rel in cell_block['Relationships']:
                                                if rel['Type'] == 'CHILD':
                                                    for word_id in rel['Ids']:
                                                        word_block = next((w for w in data['Blocks'] if w['Id'] == word_id), None)
                                                        if word_block and word_block['BlockType'] == 'WORD':
                                                            cell_text += ' ' + word_block.get('Text', '')
                                        table[row_index][col_index] = cell_text.strip()
                    table_df = pd.DataFrame.from_dict(table, orient='index').sort_index()
                    table_df = table_df.sort_index(axis=1)
                    tables.append(table_df)
            all_tables = pd.concat(tables, axis=0, ignore_index=True)
            column_a = all_tables.columns[0]

            st.subheader("Data Preview")
            st.dataframe(all_tables)

            st.subheader("Rename Columns")
            new_column_names = {}
            fiscal_year_options = [f"FY{year}" for year in range(2017, 2027)]
            ytd_options = [f"YTD{quarter}{year}" for year in range(2017, 2027) for quarter in range(1, 4)]
            dropdown_options = [''] + ['Account'] + fiscal_year_options + ytd_options

            for col in all_tables.columns:
                new_name_text = st.text_input(f"Rename '{col}' to:", value=col, key=f"rename_{col}_text")
                new_name_dropdown = st.selectbox(f"Or select predefined name for '{col}':", dropdown_options, key=f"rename_{col}_dropdown")
                new_column_names[col] = new_name_dropdown if new_name_dropdown else new_name_text
            
            all_tables.rename(columns=new_column_names, inplace=True)
            st.write("Updated Columns:", all_tables.columns.tolist())
            st.dataframe(all_tables)

            st.subheader("Edit and Remove Rows")
            editable_df = st.experimental_data_editor(all_tables, num_rows="dynamic", use_container_width=True)

            st.subheader("Select columns to keep before export")
            columns_to_keep = []
            for col in editable_df.columns:
                if st.checkbox(f"Keep column '{col}'", value=True, key=f"keep_{col}_tab1"):
                    columns_to_keep.append(col)

            editable_df = editable_df[columns_to_keep]

            st.subheader("Select numerical columns")
            numerical_columns = []
            for col in editable_df.columns:
                if st.checkbox(f"Numerical column '{col}'", value=False, key=f"num_{col}"):
                    numerical_columns.append(col)

            st.subheader("Convert Units")
            selected_columns = st.multiselect("Select columns for conversion", options=numerical_columns, key="columns_selection")
            selected_conversion_factor = st.radio("Select conversion factor", options=list(conversion_factors.keys()), key="conversion_factor")

            if st.button("Apply Selected Labels and Generate Excel", key="apply_selected_labels_generate_excel_tab1"):
                updated_table = editable_df

                for col in numerical_columns:
                    updated_table[col] = updated_table[col].apply(clean_numeric_value_IS)
                
                if selected_conversion_factor and selected_conversion_factor in conversion_factors:
                    conversion_factor = conversion_factors[selected_conversion_factor]
                    updated_table = apply_unit_conversion_IS(updated_table, selected_columns, conversion_factor)

                updated_table.replace('-', 0, inplace=True)

                excel_file = io.BytesIO()
                updated_table.to_excel(excel_file, index=False)
                excel_file.seek(0)
                st.download_button("Download Excel", excel_file, "Table_Extractor_Income_Statement.xlsx", "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

    with tab2:
        st.subheader("Aggregate My Data")

        uploaded_files = st.file_uploader("Upload Excel files", type=['xlsx'], accept_multiple_files=True, key='excel_uploader_amd')
        if uploaded_files:
            aggregated_df = aggregate_data_IS(uploaded_files)
            if aggregated_df is not None:
                st.subheader("Aggregated Data Preview")

                st.write("Select and order columns")
                columns_to_order = st.multiselect("Select columns in order", aggregated_df.columns.tolist(), default=aggregated_df.columns.tolist())
                
                if columns_to_order:
                    aggregated_df = aggregated_df[columns_to_order]

                editable_df = st.experimental_data_editor(aggregated_df, use_container_width=True)
                editable_df_excluded = editable_df.iloc[:-1]

                for col in editable_df_excluded.columns:
                    if col not in ['Account', 'Sort Index', 'Positive Decreases NI']:
                        editable_df_excluded[col] = pd.to_numeric(editable_df_excluded[col], errors='coerce').fillna(0)

                numeric_cols = editable_df_excluded.select_dtypes(include='number').columns.tolist()
                for index, row in editable_df_excluded.iterrows():
                    if row['Positive Decreases NI'] and row['Sort Index'] != 100:
                        for col in numeric_cols:
                            if col not in ['Sort Index']:
                                editable_df_excluded.at[index, col] = row[col] * -1

                final_df = pd.concat([editable_df_excluded, editable_df.iloc[-1:]], ignore_index=True)

                if 'Positive Decreases NI' in final_df.columns:
                    final_df.drop(columns=['Positive Decreases NI'], inplace=True)

                excel_file = io.BytesIO()
                final_df.to_excel(excel_file, index=False)
                excel_file.seek(0)
                st.download_button("Download Excel", excel_file, "Aggregate_My_Data_Income_Statement.xlsx", "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

    with tab3:
        st.subheader("Mappings and Data Consolidation")

        uploaded_excel_is = st.file_uploader("Upload your Excel file for Mnemonic Mapping", type=['xlsx'], key='excel_uploader_tab3_is')

        currency_options_is = ["U.S. Dollar", "Euro", "British Pound Sterling", "Japanese Yen"]
        magnitude_options_is = ["Actuals", "Thousands", "Millions", "Billions", "Trillions"]

        selected_currency_is = st.selectbox("Select Currency", currency_options_is, key='currency_selection_tab3_is')
        selected_magnitude_is = st.selectbox("Select Magnitude", magnitude_options_is, key='magnitude_selection_tab3_is')
        company_name_is = st.text_input("Enter Company Name", key='company_name_input_is')

        statement_dates = {}
        if uploaded_excel_is is not None:
            df_is = pd.read_excel(uploaded_excel_is)
            
            for col in df_is.columns:
                if col not in ['Account', 'Mnemonic', 'Manual Selection', 'Sort Index']:
                    statement_dates[col] = st.text_input(f"Enter statement date for {col}", key=f"statement_date_{col}")

            st.write("Columns in the uploaded file:", df_is.columns.tolist())

            if 'Account' not in df_is.columns:
                st.error("The uploaded file does not contain an 'Account' column.")
            else:
                if 'Sort Index' not in df_is.columns:
                    df_is['Sort Index'] = range(1, len(df_is) + 1)

                def get_best_match_is(account):
                    best_score_is = float('inf')
                    best_match_is = None
                    for _, lookup_row in income_statement_lookup_df.iterrows():
                        lookup_account = lookup_row['Account']
                        account_str = str(account)
                        score_is = levenshtein_distance(account_str.lower(), lookup_account.lower()) / max(len(account_str), len(lookup_account))
                        if score_is < best_score_is:
                            best_score_is = score_is
                            best_match_is = lookup_row
                    return best_match_is, best_score_is

                df_is['Mnemonic'] = ''
                df_is['Manual Selection'] = ''
                for idx, row in df_is.iterrows():
                    account_value = row['Account']
                    if pd.notna(account_value):
                        best_match_is, score_is = get_best_match_is(account_value)
                        if best_match_is is not None and score_is < 0.25:
                            df_is.at[idx, 'Mnemonic'] = best_match_is['Mnemonic']
                        else:
                            df_is.at[idx, 'Mnemonic'] = 'Human Intervention Required'
                    
                    if df_is.at[idx, 'Mnemonic'] == 'Human Intervention Required':
                        message_is = f"**Human Intervention Required for:** {account_value} - Index {idx}"
                        st.markdown(message_is)
                    
                    unique_mappings = income_statement_lookup_df['Mnemonic'].drop_duplicates().tolist()
                    manual_selection_is = st.selectbox(
                        f"Select category for '{account_value}'",
                        options=[''] + unique_mappings + ['REMOVE ROW'],
                        key=f"select_{idx}_tab3_is"
                    )
                    if manual_selection_is:
                        df_is.at[idx, 'Manual Selection'] = manual_selection_is.strip()

                st.dataframe(df_is[['Account', 'Mnemonic', 'Manual Selection', 'Sort Index']])

                if st.button("Generate Excel with Lookup Results", key="generate_excel_lookup_results_tab3_is"):
                    df_is['Final Mnemonic Selection'] = df_is.apply(
                        lambda row: row['Manual Selection'] if row['Manual Selection'] not in ['REMOVE ROW', ''] else row['Mnemonic'], 
                        axis=1
                    )
                    final_output_df_is = df_is[df_is['Final Mnemonic Selection'].str.strip() != 'REMOVE ROW'].copy()
                    
                    combined_df_is = create_combined_df_IS([final_output_df_is])
                    combined_df_is = sort_by_sort_index(combined_df_is)

                    def lookup_ciq_is(mnemonic):
                        if mnemonic == 'Human Intervention Required':
                            return 'CIQ ID Required'
                        ciq_value_is = income_statement_lookup_df.loc[income_statement_lookup_df['Mnemonic'] == mnemonic, 'CIQ']
                        if ciq_value_is.empty:
                            return 'CIQ ID Required'
                        return ciq_value_is.values[0]
                    
                    combined_df_is['CIQ'] = combined_df_is['Final Mnemonic Selection'].apply(lookup_ciq_is)

                    columns_order_is = ['Final Mnemonic Selection', 'CIQ'] + [col for col in combined_df_is.columns if col not in ['Final Mnemonic Selection', 'CIQ']]
                    combined_df_is = combined_df_is[columns_order_is]

                    # Update negative values just before generating the Excel file
                    combined_df_is = update_negative_values(combined_df_is)

                    as_presented_df_is = final_output_df_is.drop(columns=['CIQ', 'Mnemonic', 'Manual Selection'], errors='ignore')
                    as_presented_df_is = sort_by_sort_index(as_presented_df_is)
                    as_presented_df_is = as_presented_df_is.drop(columns=['Sort Index'], errors='ignore')
                    as_presented_columns_order_is = ['Account', 'Final Mnemonic Selection'] + [col for col in as_presented_df_is.columns if col not in ['Account', 'Final Mnemonic Selection']]
                    as_presented_df_is = as_presented_df_is[as_presented_columns_order_is]

                    excel_file_is = io.BytesIO()
                    with pd.ExcelWriter(excel_file_is, engine='xlsxwriter') as writer:
                        combined_df_is.to_excel(writer, sheet_name='Standardized - Income Stmt', index=False)
                        as_presented_df_is.to_excel(writer, sheet_name='As Presented - Income Stmt', index=False)
                        cover_df_is = pd.DataFrame({
                            'Selection': ['Currency', 'Magnitude', 'Company Name'] + list(statement_dates.keys()),
                            'Value': [selected_currency_is, selected_magnitude_is, company_name_is] + list(statement_dates.values())
                        })
                        cover_df_is.to_excel(writer, sheet_name='Cover', index=False)
                    excel_file_is.seek(0)
                    st.download_button("Download Excel", excel_file_is, "Mappings_and_Data_Consolidation_Income_Statement.xlsx", "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

                if st.button("Update Data Dictionary with Manual Mappings", key="update_data_dictionary_tab3_is"):
                    df_is['Final Mnemonic Selection'] = df_is.apply(
                        lambda row: row['Manual Selection'] if row['Manual Selection'] not in ['REMOVE ROW', ''] else row['Mnemonic'], 
                        axis=1
                    )
                    new_entries_is = []
                    for idx, row in df_is.iterrows():
                        manual_selection_is = row['Manual Selection']
                        final_mnemonic_is = row['Final Mnemonic Selection']
                        if manual_selection_is == 'REMOVE ROW':
                            continue
                        ciq_value_is = income_statement_lookup_df.loc[income_statement_lookup_df['Mnemonic'] == final_mnemonic_is, 'CIQ'].values[0] if not income_statement_lookup_df.loc[income_statement_lookup_df['Mnemonic'] == final_mnemonic_is, 'CIQ'].empty else 'CIQ ID Required'
                        
                        if manual_selection_is not in ['REMOVE ROW', '']:
                            if row['Account'] not in income_statement_lookup_df['Account'].values:
                                new_entries_is.append({'Account': row['Account'], 'Mnemonic': final_mnemonic_is, 'CIQ': ciq_value_is})
                            else:
                                income_statement_lookup_df.loc[income_statement_lookup_df['Account'] == row['Account'], 'Mnemonic'] = final_mnemonic_is
                                income_statement_lookup_df.loc[income_statement_lookup_df['Account'] == row['Account'], 'CIQ'] = ciq_value_is
                    if new_entries_is:
                        income_statement_lookup_df = pd.concat([income_statement_lookup_df, pd.DataFrame(new_entries_is)], ignore_index=True)
                    income_statement_lookup_df.reset_index(drop=True, inplace=True)
                    save_lookup_table(income_statement_lookup_df, income_statement_data_dictionary_file)
                    st.success("Data Dictionary Updated Successfully")

    with tab4:
        st.subheader("Income Statement Data Dictionary")

        if 'income_statement_data' not in st.session_state:
            if os.path.exists(income_statement_data_dictionary_file):
                st.session_state.income_statement_data = pd.read_excel(income_statement_data_dictionary_file)
            else:
                st.session_state.income_statement_data = pd.DataFrame()

        uploaded_dict_file_is = st.file_uploader("Upload a new Data Dictionary CSV", type=['csv'], key='dict_uploader_tab4_is')
        if uploaded_dict_file_is is not None:
            new_lookup_df_is = pd.read_csv(uploaded_dict_file_is)
            st.session_state.income_statement_data = new_lookup_df_is
            save_lookup_table(new_lookup_df_is, income_statement_data_dictionary_file)
            st.success("Data Dictionary uploaded and updated successfully!")

        st.dataframe(st.session_state.income_statement_data)

        remove_indices_is = st.multiselect("Select rows to remove", st.session_state.income_statement_data.index, key='remove_indices_tab4_is')
        if st.button("Remove Selected Rows", key="remove_selected_rows_tab4_is"):
            st.session_state.income_statement_data = st.session_state.income_statement_data.drop(remove_indices_is).reset_index(drop=True)
            save_lookup_table(st.session_state.income_statement_data, income_statement_data_dictionary_file)
            st.success("Selected rows removed successfully!")
            st.dataframe(st.session_state.income_statement_data)

        if st.button("Download Data Dictionary", key="download_data_dictionary_tab4_is"):
            excel_file_is = io.BytesIO()
            st.session_state.income_statement_data.to_excel(excel_file_is, index=False)
            excel_file_is.seek(0)
            st.download_button("Download Excel", excel_file_is, "income_statement_data_dictionary.xlsx", "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

                               
####################################### Populate CIQ Template ###################################
import streamlit as st
import pandas as pd
import openpyxl
from openpyxl import load_workbook
from io import BytesIO

def populate_ciq_template_pt():
    st.title("Populate CIQ Template")

    tab1, tab2 = st.tabs(["Annual Upload Template", "Quarterly Upload Template"])

    def process_template(template_type):
        unique_id = template_type.lower()

        uploaded_template = st.file_uploader(f"Upload CIQ Template ({template_type})", type=['xlsx', 'xlsm'], key=f"template_{unique_id}")
        uploaded_balance_sheet = st.file_uploader(f"Upload Completed Balance Sheet Data ({template_type})", type=['xlsx', 'xlsm'], key=f"balance_sheet_{unique_id}")
        uploaded_cash_flow = st.file_uploader(f"Upload Completed Cash Flow Statement ({template_type})", type=['xlsx', 'xlsm'], key=f"cash_flow_{unique_id}")
        uploaded_income_statement = st.file_uploader(f"Upload Completed Income Statement Data ({template_type})", type=['xlsx', 'xlsm'], key=f"income_statement_{unique_id}")

        if st.button(f"Populate {template_type} Template Now", key=f"populate_button_{unique_id}") and uploaded_template and (uploaded_balance_sheet or uploaded_cash_flow or uploaded_income_statement):
            try:
                # Read the uploaded template file
                template_file = uploaded_template.read()
                try:
                    template_wb = load_workbook(BytesIO(template_file), keep_vba=True)
                except Exception as e:
                    st.error(f"Error loading template file: {e}")
                    return

                def process_sheet(sheet_file, sheet_name, row_range, date_row):
                    try:
                        sheet_wb = load_workbook(BytesIO(sheet_file), keep_vba=True)
                    except Exception as e:
                        st.error(f"Error loading {sheet_name} file: {e}")
                        return
                    
                    as_presented_sheet = sheet_wb[f"As Presented - {sheet_name}"]
                    standardized_sheet = pd.read_excel(BytesIO(sheet_file), sheet_name=f"Standardized - {sheet_name}")

                    # Check for required columns in the Standardized sheet
                    if 'CIQ' not in standardized_sheet.columns:
                        st.error(f"The column 'CIQ' is missing from the Standardized - {sheet_name}.")
                        return

                    st.write(f"{sheet_name} CIQ column found, proceeding...")
                    st.write(standardized_sheet.head())

                    # Copy the "As Presented" sheet to the template workbook
                    if f"As Presented - {sheet_name}" in template_wb.sheetnames:
                        del template_wb[f"As Presented - {sheet_name}"]
                    new_sheet = template_wb.create_sheet(f"As Presented - {sheet_name}")
                    for row in as_presented_sheet.iter_rows():
                        for cell in row:
                            new_sheet[cell.coordinate].value = cell.value

                    # Color the "As Presented" tab orange
                    tab_color = "FFA500"
                    new_sheet.sheet_properties.tabColor = tab_color

                    # Copy the "Standardized" sheet to the template workbook
                    if f"Standardized - {sheet_name}" in template_wb.sheetnames:
                        del template_wb[f"Standardized - {sheet_name}"]
                    standardized_ws = template_wb.create_sheet(f"Standardized - {sheet_name}")

                    # Write the header row
                    for col_num, header in enumerate(standardized_sheet.columns, 1):
                        standardized_ws.cell(row=1, column=col_num, value=header)

                    # Write the data rows
                    for r_idx, row in enumerate(standardized_sheet.itertuples(index=False), 2):
                        for c_idx, value in enumerate(row, 1):
                            standardized_ws.cell(row=r_idx, column=c_idx, value=value)

                    # Perform lookups and update the "Upload" sheet
                    upload_sheet = template_wb["Upload"]
                    ciq_values = standardized_sheet['CIQ'].tolist()
                    dates = list(standardized_sheet.columns[1:])  # Assumes dates start from the second column

                    st.write(f"CIQ Values from {sheet_name}:", ciq_values)
                    st.write(f"Dates from {sheet_name}:", dates)

                    if template_type == "Annual":
                        for row in upload_sheet.iter_rows(min_row=row_range[0], max_row=row_range[1], min_col=4, max_col=upload_sheet.max_column):
                            ciq_cell = upload_sheet.cell(row=row[0].row, column=11)
                            ciq_value = ciq_cell.value
                            if ciq_value in ciq_values:
                                st.write(f"Processing CIQ Value: {ciq_value} at row {row[0].row}")
                                for col in range(4, 10):
                                    date_value = upload_sheet.cell(row=date_row, column=col).value
                                    st.write(f"Checking date {date_value} at column {col}")
                                    if date_value in dates:
                                        lookup_value = standardized_sheet.loc[standardized_sheet['CIQ'] == ciq_value, date_value].sum()
                                        st.write(f"Lookup value for CIQ {ciq_value} and date {date_value}: {lookup_value}")
                                        if not pd.isna(lookup_value):
                                            cell_to_update = upload_sheet.cell(row=row[0].row, column=col)
                                            if cell_to_update.data_type == 'f' or cell_to_update.value is None:
                                                cell_to_update.value = lookup_value
                                                st.write(f"Updated {cell_to_update.coordinate} with value {lookup_value}")

                        for row in upload_sheet.iter_rows(min_row=row_range[1] + 1, max_row=row_range[1] + 1, min_col=4, max_col=9):
                            for cell in row:
                                if cell.value is not None:
                                    try:
                                        cell_value = float(cell.value)
                                        cell.value = -abs(cell_value)
                                    except ValueError:
                                        st.warning(f"Non-numeric value found in cell {cell.coordinate}, skipping negation.")
                    
                    elif template_type == "Quarterly":
                        for row in upload_sheet.iter_rows(min_row=row_range[0], max_row=row_range[1], min_col=4, max_col=21):  # Changed to column U (21)
                            ciq_cell = upload_sheet.cell(row=row[0].row, column=23)  # Changed to column W (23)
                            ciq_value = ciq_cell.value
                            if ciq_value in ciq_values:
                                st.write(f"Processing CIQ Value: {ciq_value} at row {row[0].row}")
                                for col in range(4, 22):  # Changed to include columns D to U
                                    date_value = upload_sheet.cell(row=date_row, column=col).value
                                    st.write(f"Checking date {date_value} at column {col}")
                                    if date_value in dates:
                                        lookup_value = standardized_sheet.loc[standardized_sheet['CIQ'] == ciq_value, date_value].sum()
                                        st.write(f"Lookup value for CIQ {ciq_value} and date {date_value}: {lookup_value}")
                                        if not pd.isna(lookup_value):
                                            cell_to_update = upload_sheet.cell(row=row[0].row, column=col)
                                            if cell_to_update.data_type == 'f' or cell_to_update.value is None:
                                                cell_to_update.value = lookup_value
                                                st.write(f"Updated {cell_to_update.coordinate} with value {lookup_value}")

                        for row in upload_sheet.iter_rows(min_row=row_range[1] + 1, max_row=row_range[1] + 1, min_col=4, max_col=21):  # Changed to column U (21)
                            for cell in row:
                                if cell.value is not None:
                                    try:
                                        cell_value = float(cell.value)
                                        cell.value = -abs(cell_value)
                                    except ValueError:
                                        st.warning(f"Non-numeric value found in cell {cell.coordinate}, skipping negation.")

                # Process sheets based on the uploaded files
                if template_type == "Annual":
                    if uploaded_balance_sheet:
                        balance_sheet_file = uploaded_balance_sheet.read()
                        process_sheet(balance_sheet_file, "Balance Sheet", (94, 160), 92)
                    if uploaded_cash_flow:
                        cash_flow_file = uploaded_cash_flow.read()
                        process_sheet(cash_flow_file, "Cash Flow", (169, 232), 167)
                    if uploaded_income_statement:
                        income_statement_file = uploaded_income_statement.read()
                        process_sheet(income_statement_file, "Income Stmt", (12, 70), 10)

                if template_type == "Quarterly":
                    if uploaded_balance_sheet:
                        balance_sheet_file = uploaded_balance_sheet.read()
                        process_sheet(balance_sheet_file, "Balance Sheet", (94, 160), 92)
                    if uploaded_cash_flow:
                        cash_flow_file = uploaded_cash_flow.read()
                        process_sheet(cash_flow_file, "Cash Flow", (169, 232), 167)
                    if uploaded_income_statement:
                        income_statement_file = uploaded_income_statement.read()
                        process_sheet(income_statement_file, "Income Stmt", (12, 70), 10)

                # Save the updated workbook to a BytesIO object
                output = BytesIO()
                template_wb.save(output)
                template_data = output.getvalue()

                # Provide a download button for the updated template
                st.download_button(
                    label=f"Download Updated {template_type} Template",
                    data=template_data,
                    file_name=uploaded_template.name,
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                )

                st.success(f"{template_type} Template populated successfully. You can now download the updated template.")

            except Exception as e:
                st.error(f"An error occurred: {e}")

    with tab1:
        process_template("Annual")

    with tab2:
        process_template("Quarterly")
#######################################Extras#############################
import io
import os
import json
import pandas as pd
import streamlit as st
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from Levenshtein import distance as levenshtein_distance
import re

# Ensure other functions and imports remain unchanged

# New function to handle the Extras tab
def extras_tab():
    st.title("Extras")

    if st.button("Backup Data Dictionaries"):
        # Load data dictionaries
        balance_sheet_data = pd.read_csv('balance_sheet_data_dictionary.csv')
        cash_flow_data = pd.read_csv('cash_flow_data_dictionary.csv')
        income_statement_data = pd.read_excel('income_statement_data_dictionary.xlsx')

        # Create a new Excel writer object
        with pd.ExcelWriter("data_dictionaries_backup.xlsx", engine='xlsxwriter') as writer:
            # Write each DataFrame to a different sheet
            balance_sheet_data.to_excel(writer, sheet_name='Balance Sheet', index=False)
            cash_flow_data.to_excel(writer, sheet_name='Cash Flow', index=False)
            income_statement_data.to_excel(writer, sheet_name='Income Statement', index=False)

        # Read the file into a BytesIO object for download
        with open("data_dictionaries_backup.xlsx", "rb") as file:
            backup_file = io.BytesIO(file.read())

        st.download_button(
            label="Download Backup",
            data=backup_file,
            file_name="data_dictionaries_backup.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )

def main():
    st.sidebar.title("Navigation")
    selection = st.sidebar.radio("Go to", ["Balance Sheet", "Cash Flow Statement", "Income Statement", "Populate CIQ Template", "Extras"])

    if selection == "Balance Sheet":
        balance_sheet()
    elif selection == "Cash Flow Statement":
        cash_flow_statement()
    elif selection == "Income Statement":
        income_statement()
    elif selection == "Populate CIQ Template":
        populate_ciq_template_pt()
    elif selection == "Extras":
        extras_tab()

if __name__ == '__main__':
    main()


# In[ ]:




